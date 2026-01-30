"""
Excel handler for reading and writing invoice data
"""

import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
from config import TEMPLATE_FILE, INVOICE_FIELDS, OUTPUT_FOLDER


class ExcelHandler:
    """Handle Excel operations for invoice template"""
    
    def __init__(self, template_path=TEMPLATE_FILE):
        """Initialize with template path"""
        self.template_path = template_path
        self.workbook = None
        self.worksheet = None
        
    def load_template(self):
        """Load the template Excel file"""
        try:
            if not os.path.exists(self.template_path):
                raise FileNotFoundError(f"Template file not found: {self.template_path}")
            
            self.workbook = load_workbook(self.template_path)
            # Get the first sheet or 'Invoice' sheet
            sheet_name = 'Invoice' if 'Invoice' in self.workbook.sheetnames else self.workbook.sheetnames[0]
            self.worksheet = self.workbook[sheet_name]
            return True
        except Exception as e:
            raise Exception(f"Error loading template: {str(e)}")
    
    def get_cell_value(self, cell_ref):
        """Get value from a specific cell"""
        try:
            # support list of cells
            if isinstance(cell_ref, (list, tuple)):
                values = []
                for c in cell_ref:
                    values.append(self.worksheet[c].value if self.worksheet[c].value is not None else '')
                # join with newline for multi-line fields
                return "\n".join(str(v) for v in values).strip()
            return self.worksheet[cell_ref].value
        except Exception as e:
            raise Exception(f"Error reading cell {cell_ref}: {str(e)}")
    
    def set_cell_value(self, cell_ref, value):
        """Set value to a specific cell"""
        try:
            # support list of cells
            if isinstance(cell_ref, (list, tuple)):
                # If value is a string with newlines, split into lines
                if isinstance(value, str) and "\n" in value:
                    parts = value.splitlines()
                # If value is a list/tuple, use it directly
                elif isinstance(value, (list, tuple)):
                    parts = [str(v) for v in value]
                else:
                    # single scalar: write the same value to all target cells
                    parts = [value]

                # write parts into cells in order; if fewer parts than cells, fill remaining with empty string
                for idx, c in enumerate(cell_ref):
                    v = parts[idx] if idx < len(parts) else ''
                    self.worksheet[c].value = v
                return
            self.worksheet[cell_ref].value = value
        except Exception as e:
            raise Exception(f"Error writing to cell {cell_ref}: {str(e)}")
    
    def get_all_template_values(self):
        """Get all template field values"""
        values = {}
        try:
            for field_key, field_config in INVOICE_FIELDS.items():
                cell_ref = field_config['cell']
                try:
                    values[field_key] = self.get_cell_value(cell_ref)
                except:
                    values[field_key] = ''
            return values
        except Exception as e:
            raise Exception(f"Error reading template values: {str(e)}")
    
    def update_invoice(self, data_dict):
        """Update invoice with provided data"""
        try:
            for field_key, value in data_dict.items():
                if field_key in INVOICE_FIELDS:
                    field_config = INVOICE_FIELDS[field_key]
                    if not field_config.get('read_only', False):
                        cell_ref = field_config['cell']
                        # For date fields that may be strings, try to keep them as-is; the template will display string
                        self.set_cell_value(cell_ref, value)
        except Exception as e:
            raise Exception(f"Error updating invoice: {str(e)}")
    
    def save_invoice(self, output_filename=None):
        """Save the modified invoice"""
        try:
            # Create output folder if it doesn't exist
            if not os.path.exists(OUTPUT_FOLDER):
                os.makedirs(OUTPUT_FOLDER)
            
            if output_filename is None:
                # Try to use invoice_no from template as filename if present
                try:
                    inv_field = INVOICE_FIELDS.get('invoice_no', {})
                    inv_cell = inv_field.get('cell')
                    inv_val = None
                    if inv_cell:
                        inv_val = self.get_cell_value(inv_cell)
                    if inv_val:
                        safe_name = str(inv_val).strip().replace('/', '-').replace('\\\n', '_')
                        output_filename = f"{safe_name}.xlsx"
                    else:
                        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                        output_filename = f"Invoice_{timestamp}.xlsx"
                except Exception:
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    output_filename = f"Invoice_{timestamp}.xlsx"
            
            output_path = os.path.join(OUTPUT_FOLDER, output_filename)
            self.workbook.save(output_path)
            return output_path
        except Exception as e:
            raise Exception(f"Error saving invoice: {str(e)}")
    
    def close(self):
        """Close the workbook"""
        if self.workbook:
            self.workbook.close()
